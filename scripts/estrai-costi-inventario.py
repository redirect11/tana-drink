# =====================================================================
#  Estrae i COSTI AL DETTAGLIO dei prodotti dal GENERATORE ORDINI (o da
#  INV.xlsx). Colonne usate su ogni foglio: "art" (nome), "cl" (contenuto
#  della confezione), "TIPO" (categoria), "€/pz" (costo NETTO per
#  confezione). Da questi si ricava il costo al cl come nell'Excel:
#      €/cl = (€/pz + IVA) / cl
#
#    python scripts/estrai-costi-inventario.py [--file "GEN ORD REC.xlsx"] [--sheet NOME]
#
#  Il generatore ordini ha un foglio per ordine, in ordine cronologico:
#  senza --sheet si scorrono TUTTI e per ogni prodotto si tiene il PREZZO
#  PIÙ RECENTE (l'ultima occorrenza vince). Con --sheet si legge solo quello.
#
#  Scrive costi-inventario.json (dati aziendali: gitignorato). Poi:
#      node scripts/import-costi-inventario.js            # anteprima
#      node scripts/import-costi-inventario.js --apply    # scrive
# =====================================================================
import argparse
import json
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('Serve openpyxl: pip install openpyxl')

ap = argparse.ArgumentParser()
ap.add_argument('--file', default='GEN ORD REC.xlsx')
ap.add_argument('--sheet', default=None, help='un solo foglio (default: tutti, più recente vince)')
ap.add_argument('--out', default='costi-inventario.json')
# L'ASSORTIMENTO (linea / premium / fuori) sta nel generatore ordini, non in
# INV: la colonna con LINEA, PREM e OUT e' li'. Si legge da un file a parte
# e si unisce ai prodotti per nome.
ap.add_argument('--stati-da', default='GEN ORD REC.xlsx',
                help='file da cui leggere LINEA/PREM/OUT (vuoto per saltare)')
ap.add_argument('--vat', type=float, default=22.0, help='aliquota IVA usata nel foglio')
ap.add_argument('--max-rows', type=int, default=1300, help='righe max per foglio (evita i fogli vuoti enormi)')
args = ap.parse_args()

wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def find_header(ws):
    """Riga di intestazione (contiene 'art') e indici colonna."""
    for i, row in enumerate(ws.iter_rows(max_row=20, max_col=24, values_only=True), start=1):
        valori = [(j, str(c).strip().lower()) for j, c in enumerate(row) if c]
        if any(v == 'art' for _, v in valori):
            cols = {}
            for j, v in valori:
                if v == 'art':
                    cols['name'] = j
                elif v == 'cl':
                    cols['cl'] = j
                elif v == 'tipo':
                    cols['tipo'] = j
                elif v.startswith('€/pz') or v.endswith('/pz'):
                    cols['cost'] = j
                # DEP = deposito, cioè la GIACENZA in confezioni (0.8 = 8/10
                # di bottiglia). Serve a riallineare il magazzino, non solo
                # il costo.
                elif v == 'dep':
                    cols['dep'] = j
                # Colonna "C": porta LINEA (i prodotti che non devono
                # mancare) e OUT (fuori assortimento).
                elif v == 'c':
                    cols['stato'] = j
            if 'name' in cols:
                return i, cols
    return None, None


def read_sheet(ws, into):
    """Legge un foglio e aggiorna `into` (nome normalizzato → prodotto):
    l'ultima occorrenza (foglio più recente) sovrascrive."""
    header_row, cols = find_header(ws)
    if header_row is None or 'cost' not in cols:
        return 0
    added = 0
    for k, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_col=24, values_only=True)):
        if k > args.max_rows:
            break
        nome = row[cols['name']]
        if not isinstance(nome, str) or not nome.strip():
            continue
        costo = num(row[cols['cost']])
        if costo is None or costo <= 0:
            continue
        cl = num(row[cols['cl']]) if 'cl' in cols else None
        dep = num(row[cols['dep']]) if 'dep' in cols else None
        stato = None
        if 'stato' in cols and row[cols['stato']]:
            stato = str(row[cols['stato']]).strip().upper() or None
        nome = nome.strip()
        precedente = into.get(nome.lower(), {})
        into[nome.lower()] = {
            'name': nome,
            'cl': cl,
            'tipo': (row[cols['tipo']] or '').strip() if 'tipo' in cols and row[cols['tipo']] else None,
            'cost': round(costo, 4),  # NETTO per confezione
            'vat': args.vat,
            'package_size_ml': int(round(cl * 10)) if cl else None,
            'cost_per_cl': round(costo * (1 + args.vat / 100) / cl, 6) if cl else None,
            # Giacenza in CONFEZIONI dall'ultimo foglio che nomina il prodotto.
            'dep': round(dep, 4) if dep is not None and dep >= 0 else None,
            # OUT vale l'ULTIMO foglio (è la situazione di adesso); LINEA
            # sta solo sul foglio ORD e non va persa dai fogli successivi,
            # che quella colonna non la compilano.
            'out': stato == 'OUT',
            'linea': stato == 'LINEA' or bool(precedente.get('linea')),
        }
        added += 1
    return added


latest = {}
if args.sheet:
    if args.sheet not in wb.sheetnames:
        sys.exit(f'Foglio "{args.sheet}" assente.')
    read_sheet(wb[args.sheet], latest)
    origine = args.sheet
else:
    # Tutti i fogli in ordine: l'ultimo che nomina un prodotto ne fissa il
    # prezzo corrente (i fogli sono cronologici).
    fogli = 0
    for name in wb.sheetnames:
        if read_sheet(wb[name], latest):
            fogli += 1
    origine = f'{fogli} fogli del generatore'


# ── ASSORTIMENTO: linea / premium / fuori ────────────────────────────
# I marcatori stanno nelle prime colonne, ma non sempre nella stessa: in un
# foglio OUT sta in "X" e LINEA/PREM in "C", in un altro sono insieme. Invece
# di fidarsi della posizione si cercano i TOKEN nelle prime colonne.
TOKEN = {'OUT': 'out', 'LINEA': 'linea', 'PREM': 'premium', 'PREMIUM': 'premium'}


def leggi_stati(path):
    stati = {}
    try:
        wbs = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f'[costi] stati non letti da {path}: {e}')
        return stati
    for name in wbs.sheetnames:
        ws2 = wbs[name]
        header_row, cols2 = find_header(ws2)
        if header_row is None or 'name' not in cols2:
            continue
        jn = cols2['name']
        for k, row in enumerate(ws2.iter_rows(min_row=header_row + 1, max_col=24, values_only=True)):
            if k > args.max_rows:
                break
            nome = row[jn] if jn < len(row) else None
            if not isinstance(nome, str) or not nome.strip():
                continue
            trovato = None
            for v in row[:jn]:
                if isinstance(v, str) and v.strip().upper() in TOKEN:
                    trovato = TOKEN[v.strip().upper()]
                    break
            # L'ultimo foglio che si esprime su un prodotto ha ragione: i
            # fogli sono in ordine e il piu' recente e' la situazione di ora.
            if trovato:
                stati[nome.strip().lower()] = trovato
    return stati


if args.stati_da:
    stati = leggi_stati(args.stati_da)
    quanti = {'linea': 0, 'premium': 0, 'out': 0, 'senza prodotto': 0}
    for nome, st in stati.items():
        if nome in latest:
            latest[nome]['stato'] = st
            quanti[st] += 1
        else:
            quanti['senza prodotto'] += 1
    print(
        f"[costi] assortimento da {args.stati_da}: "
        f"{quanti['linea']} in linea, {quanti['premium']} premium, {quanti['out']} fuori"
        + (f" ({quanti['senza prodotto']} nomi non nel listino)" if quanti['senza prodotto'] else '')
    )

prodotti = sorted(latest.values(), key=lambda p: p['name'].lower())
with open(args.out, 'w', encoding='utf-8') as f:
    json.dump({'sheet': origine, 'vat': args.vat, 'products': prodotti}, f, ensure_ascii=False, indent=2)

con_dep = sum(1 for p in prodotti if p.get('dep') is not None)
in_linea = sum(1 for p in prodotti if p.get('linea'))
fuori = sum(1 for p in prodotti if p.get('out'))
print(f'[costi] {origine}: {len(prodotti)} prodotti con prezzo più recente.')
print(f'[costi] {con_dep} con la giacenza (colonna DEP).')
print(f'[costi] assortimento (colonna C): {in_linea} in linea, {fuori} fuori.')
print(f'[costi] scritto {args.out}')
